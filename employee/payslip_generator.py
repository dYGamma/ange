# employee/payslip_generator.py

import os
import sys
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font
from plyer import notification
from PyQt5.QtWidgets import QFileDialog

class PayslipGenerator:
    PDF_FONT_NAME = "SystemArial"

    def __init__(self, out_dir="payslips"):
        os.makedirs(out_dir, exist_ok=True)
        self.out_dir = out_dir

        # Найти и зарегистрировать Arial.ttf
        font_path = None
        if sys.platform.startswith("win"):
            font_path = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arial.ttf")
        elif sys.platform == "darwin":
            font_path = "/Library/Fonts/Arial.ttf"
        else:
            for p in (
                "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
                "/usr/share/fonts/truetype/msttcorefonts/arial.ttf",
                "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            ):
                if os.path.isfile(p):
                    font_path = p
                    break

        if not font_path or not os.path.isfile(font_path):
            raise FileNotFoundError(f"Не найден системный TTF‑шрифт Arial по пути: {font_path}")

        pdfmetrics.registerFont(TTFont(self.PDF_FONT_NAME, font_path))

    def _get_save_path(self, parent, default_name, filter):
        """
        Открывает диалог «Сохранить как» и возвращает полный путь.
        Если пользователь отменил — возвращает None.
        """
        path, _ = QFileDialog.getSaveFileName(parent,
                                              "Сохранить файл",
                                              os.path.join(self.out_dir, default_name),
                                              filter)
        return path or None

    def generate_pdf(self, user, salary_record, parent=None):
        # Предложить пользователю ввести имя и путь файла
        default_name = f"payslip_{user.id}_{salary_record.month}.pdf"
        save_path = self._get_save_path(parent, default_name, "PDF files (*.pdf)")
        if not save_path:
            return  # пользователь отменил

        # Данные для таблички
        data = [
            ["Параметр", "Значение"],
            ["Месяц", salary_record.month],
            ["Сотрудник", f"{user.full_name} ({user.position})"],
            ["Премии", salary_record.bonus],
            ["Начислено", salary_record.gross],
            ["НДФЛ (13%)", salary_record.ndfl],
            ["К выплате", salary_record.payout],
        ]

        # Создаём canvas
        c = canvas.Canvas(save_path, pagesize=A4)
        c.setFont(self.PDF_FONT_NAME, 12)

        # Построим таблицу с Platypus
        table = Table(data, colWidths=[150, 200])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), self.PDF_FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ]))

        # Помещаем таблицу на страницу
        w, h = A4
        table_width, table_height = table.wrap(0, 0)
        x = (w - table_width) / 2
        y = h - 100  # отступ сверху
        table.drawOn(c, x, y - table_height)

        c.showPage()
        c.save()

        notification.notify(
            title="Payslip PDF",
            message=f"PDF сохранён: {save_path}"
        )

    def generate_excel(self, user, salary_record, parent=None):
        default_name = f"payslip_{user.id}_{salary_record.month}.xlsx"
        save_path = self._get_save_path(parent, default_name, "Excel files (*.xlsx)")
        if not save_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"

        rows = [
            ["Параметр", "Значение"],
            ["Месяц", salary_record.month],
            ["Сотрудник", user.full_name],
            ["Премии", salary_record.bonus],
            ["Начислено", salary_record.gross],
            ["НДФЛ (13%)", salary_record.ndfl],
            ["К выплате", salary_record.payout],
        ]

        for r, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = Font(name="Arial", size=12)
        # Добавим границы
        from openpyxl.styles.borders import Border, Side
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        max_row = len(rows)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

        wb.save(save_path)
        notification.notify(
            title="Payslip XLSX",
            message=f"Excel сохранён: {save_path}"
        )
